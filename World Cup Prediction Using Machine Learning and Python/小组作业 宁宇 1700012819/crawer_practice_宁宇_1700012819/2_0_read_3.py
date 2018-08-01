# -*- coding:utf-8 -*-
# python2.7
import json
#fRead=open(r'D:\Code\Python\2018Summer\2018Summer\2_0.json','r')
fRead=open('2_0.json','r')
data=json.loads(fRead.read())
for line in data:
    line[0]=int(line[0])
print data[0][0], data[0][1]
print 'load finish'
'''
try:
    import openpyxl
except:
    import openpyxl
book=openpyxl.Workbook()
sheet1=book.active
sheet1.title='2016'
sheet2=book.create_sheet()
sheet2.title='2017'
sheet3=book.create_sheet()
sheet3.title='2018'
for line in data:
    if line[0]==2016:
        sheet1.append(line)
    elif line[0]==2017:
        sheet2.append(line)
    elif line[0]==2018:
        sheet3.append(line)
    else:
        print line
book.save('2_0.xlsx')
book.close()
print 'Turn to xlsx finish'
'''